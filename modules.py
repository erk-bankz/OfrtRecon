import openpyxl
import docx
from pathlib import Path
import re
from openpyxl.styles import Color, Font, colors
import win32com.client as win32

# ft_file_path = Path("c:\\Users\\ehom\\Documents\\IdeaProjects\\Python\\Projects\\asposeWordExcel\\test_source_files\\test 1\\FT\\")
# bt_file_path = Path("c:\\Users\\ehom\\Documents\\IdeaProjects\\Python\\Projects\\asposeWordExcel\\test_source_files\\test 1\\BT\\")
# xls_file_path = Path("c:\\Users\\ehom\\Documents\\IdeaProjects\\Python\\Projects\\asposeWordExcel\\test_source_files\\test 1\\Report\\")
# front_user_column = 5   # 5 for recon to ort, 8 for ort to recon
# back_user_column = 6   # 6 '', 9 ''
# hideTwoRows = True


def extractfileNameandFileLP(ft_doc):
    fileFormatList = ["-MS Word_TXLF", "-MS Word_TXLF_translate comments", "-MS Excel_TXLF", "-MS PPT_TXLF", "-MS PPT_TXLF_no notes", "-IDML_TXLF",
                      "-Source_TXLF", "-XML_Medavante_TXLF", "-Non-Parsable", "-MS Word_TXML", "-MS Word_TXML_translate comments", "-MS Excel_TXML",
                      "-MS PPT_TXML", "-MS PPT_TXML_no notes", "-IDML_TXML", "-Source_TXML", "-XML_Medavante_TXML", "-ERT_JSON"]
    f_name = ft_doc.name
    f_lp = re.findall("-[a-z][a-z]-[A-Z][A-Z]", f_name)[0]
    for pdGib in fileFormatList:
        f_name = f_name.split(pdGib)[0]
        f_name = f_name.split(".docx")[0]
    return f_name, f_lp


def findBTmatch(front_doc, front_lp, bt_file_path):
    for back_trans_file in bt_file_path.iterdir():
        if back_trans_file.suffix == ".docx":
            if front_doc in back_trans_file.name and front_lp in back_trans_file.name:
                return back_trans_file


def findXLSmatch(front_doc, front_lp, xls_file_path):
    for excel_file in xls_file_path.iterdir():
        if excel_file.suffix == ".xlsx":
            if front_lp in excel_file.name and front_doc in excel_file.name:
                return excel_file


def extract_xlsx_back_values(table, back_column):
    sheet_obj = table.active
    sheet_max_row = sheet_obj.max_row
    back_col = []
    for i in range(4, sheet_max_row+1):
        cell_obj = sheet_obj.cell(row=i, column=back_column)
        back_col.append(cell_obj.value)
    return back_col


def extract_xlsx_front_values(table, front_column):
    sheet_obj = table.active
    sheet_max_row = sheet_obj.max_row
    front_col = []
    for i in range(4, sheet_max_row+1):
        cell_obj = sheet_obj.cell(row=i, column=front_column)
        front_col.append(cell_obj.value)
    return front_col


def extract_table_values(table):
    keys = None
    data = []
    text_to_return = []
    for i, row in enumerate(table.rows):
        text = (cell.text for cell in row.cells)
        if i == 0:
            keys = tuple(text)
            continue
        row_data = dict(zip(keys, text))
        data.append(row_data)
    for row in data:
        text_to_return.append(row[keys[2]])
    return text_to_return


def ORTtoRecon(ft_ort_table,bt_ort_table,active_sheet,front_user_column,back_user_column):
    front_row_start, back_row_start = 4, 4
    for word_cell in ft_ort_table:
        active_sheet.cell(row=front_row_start, column=front_user_column).value = word_cell
        active_sheet.cell(row=front_row_start, column=front_user_column).font = Font(color="FF000000")
        front_row_start += 1
    for word_cell in bt_ort_table:
        active_sheet.cell(row=back_row_start, column=back_user_column).value = word_cell
        active_sheet.cell(row=back_row_start, column=back_user_column).font = Font(color="FF000000")
        back_row_start += 1



def RecontoORT(ft_ort_table, bt_ort_table, xls_file, front_user_column, back_user_column):
    count = 0
    xlsx_front_column = extract_xlsx_front_values(xls_file, front_user_column)
    xlsx_back_column = extract_xlsx_back_values(xls_file, back_user_column)
    for word_cell in ft_ort_table.rows:
        if "Target " in word_cell.cells[2].text:
            continue
        word_cell.cells[2].text=str(xlsx_front_column[count])
        count += 1
    count = 0
    for word_cell in bt_ort_table.rows:
        if "Target " in word_cell.cells[2].text:
            continue
        word_cell.cells[2].text=str(xlsx_back_column[count])
        count += 1


def removeCheckbox(xls_file_path):
    excel= win32.gencache.EnsureDispatch('Excel.Application')
    xlsx = Path(xls_file_path)
    for file in xlsx.iterdir():
        if file.suffix == ".xlsx":
            workbook=excel.Workbooks.Open(file)
            for shp in workbook.ActiveSheet.Shapes:
                shp.Delete()
            workbook.Save()
            workbook.Close()








